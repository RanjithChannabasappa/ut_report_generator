#   !/usr/bin/env
#   _*_ Coding: UTF-8 _*_
#   __version__ = “1.1”

"""
    The ut_report_generator.py module populates the UT execution details and
    creates an excel sheet with the file name given by the user as ".txt".

"""
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.chart import Reference, BarChart3D, BarChart
from openpyxl.chart.label import DataLabelList

# Printing a ut_report_generator version with welcome note
print(">>> Welcome to UT Report Generator !!! -- version: 1.1 <<<\n")

# Get a file name from user
input_file = input(">>> Please enter the UT text file name along with path and extension:\n").rstrip()

# Fetch the file name to be saved
if "\\" in input_file:
    file_extension = input_file.split("\\")[-1].split(".txt")[0]
    file_path = input_file.split(file_extension)[0]
else:
    file_extension = input_file.split(".txt")[0]
    file_path = os.getcwd() + "\\"

result_file = file_extension + "_results.xlsx"
save_file_path = file_path + result_file


# Creating a dictionary to save values for row headers
excel_dict = {}
# Creating a list for headers
excel_list = ['Sl No.', 'TEST CASE NAME', 'CLASS', 'EXECUTION', 'STATUS CODE', 'RESULT']

# Creating a Workbook
wb = Workbook()

# Delete the default sheet created while opening Workbook
wb.remove(wb["Sheet"])

# Creating a sheet
summary = wb.create_sheet("Summary")
ut_sheet = wb.create_sheet(file_extension + "_Results")

# Creating the color
color_header = PatternFill("solid", "3333ff", "3333ff")
color_title = PatternFill("solid", "9999ff", "9999ff")
color_total = PatternFill("solid", "99bbff", "99bbff")
color_pass = PatternFill("solid", "99ff99", "99ff99")
color_fail = PatternFill("solid", "ff9999", "ff9999")
color_error = PatternFill("solid", "ffff99", "ffff99")
color_ignored = PatternFill("solid", "d6d6c2", "d6d6c2")
color_tbd = PatternFill("solid", "99bbff", "99bbff")

color_header_ns = PatternFill("solid", "3333ff", "3333ff")
color_values_ns = PatternFill("solid", "f2f2f2", "f2f2f2")

# Creating the border
border_thick = Border(left=Side("thick"), right=Side("thick"), top=Side("thick"),
                      bottom=Side("thick"))
border_thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"),
                     bottom=Side("thin"))

# Creating the Alignment
alignment_header = Alignment(horizontal="center", vertical="center")

# Creating the font style
font_header = Font(size=12.5, bold=True)
font_title = Font(size=12, bold=True)
font_values = Font(size=11, bold=False)

# Creating the styles
style_header_s = NamedStyle(name="style_header_s")
style_header_s.fill = color_header
style_header_s.border = border_thick
style_header_s.alignment = alignment_header
style_header_s.font = font_header

style_header_ns = NamedStyle(name="style_header_ns")
style_header_ns.fill = color_header_ns
style_header_ns.border = border_thick
style_header_ns.alignment = alignment_header
style_header_ns.font = font_header

style_title = NamedStyle(name="style_title")
style_title.fill = color_title
style_title.border = border_thin
style_title.alignment = alignment_header
style_title.font = font_title

style_values_ns = NamedStyle(name="style_values_ns")
style_values_ns.fill = color_values_ns
style_values_ns.border = border_thin
style_values_ns.alignment = alignment_header
style_values_ns.font = font_values

style_pass = NamedStyle(name="style_pass")
style_pass.fill = color_pass
style_pass.border = border_thin
style_pass.alignment = alignment_header
style_pass.font = font_values

style_fail = NamedStyle(name="style_fail")
style_fail.fill = color_fail
style_fail.border = border_thin
style_fail.alignment = alignment_header
style_fail.font = font_values

style_error = NamedStyle(name="style_error")
style_error.fill = color_error
style_error.border = border_thin
style_error.alignment = alignment_header
style_error.font = font_values

style_ignored = NamedStyle(name="style_ignored")
style_ignored.fill = color_ignored
style_ignored.border = border_thin
style_ignored.alignment = alignment_header
style_ignored.font = font_values

style_tbd = NamedStyle(name="style_tbd")
style_tbd.fill = color_tbd
style_tbd.border = border_thin
style_tbd.alignment = alignment_header
style_tbd.font = font_values

style_total = NamedStyle(name="style_total")
style_total.fill = color_total
style_total.border = border_thin
style_total.alignment = alignment_header
style_total.font = font_values

# Initializing row and col value to zero
row = 1
col = 1

for n in excel_list:
    ut_sheet.cell(row, col, n,).style = style_header_ns
    col += 1

try:
    # Open an input UT Text file
    with open(input_file) as ut_file:
        ut_data = ut_file.readlines()
except PermissionError:
    print(">>>>>> File Permission Denied <<<<<<\n")
    sys.exit()
except FileNotFoundError:
    print(">>>>>> File Not Found <<<<<<\n")
    sys.exit()

for line in ut_data:
    if "test=" in line:
        test = line.split("test=")[1].split("\n")[0]
        print("test:", test)
    if "class=" in line:
        ut_class = line.split("class=")[1].split("\n")[0]
    if "INSTRUMENTATION_STATUS_CODE:" in line:
        status_code = int(line.split("INSTRUMENTATION_STATUS_CODE:")[1].split("\n")[0])
        if status_code != 1:
            if test in excel_dict:
                if ut_class in excel_dict[test]:
                    excel_dict[test][ut_class][0] += 1
                    excel_dict[test][ut_class][1].append(status_code)
                else:
                    excel_dict[test] = {ut_class: [1, [status_code]]}
            else:
                excel_dict[test] = {ut_class: [1, [status_code]]}

t_pass, t_fail, t_total, t_ignored, t_error, t_tbd, = 0, 0, 0, 0, 0, 0

for n in excel_dict:
    for m in excel_dict[n]:
        row += 1
        col = 1
        result = "TBD"
        t_total += 1

        ut_sheet.cell(row, col, row - 1).style = style_values_ns
        ut_sheet.cell(row, col + 1, n).style = style_values_ns
        ut_sheet.cell(row, col + 2, m).style = style_values_ns
        ut_sheet.cell(row, col + 3, excel_dict[n][m][0]).style = style_values_ns
        ut_sheet.cell(row, col + 4, str(excel_dict[n][m][1])).style = style_values_ns
        if "-3" in str(excel_dict[n][m][1]):
            result = "IGNORED"
            ut_sheet.cell(row, col + 5, result).style = style_ignored
        if "0" in str(excel_dict[n][m][1]):
            result = "PASS"
            ut_sheet.cell(row, col + 5, result).style = style_pass
        if "-1" in str(excel_dict[n][m][1]):
            result = "ERROR"
            ut_sheet.cell(row, col + 5, result).style = style_error
        if "-2" in str(excel_dict[n][m][1]):
            result = "FAIL"
            ut_sheet.cell(row, col + 5, result).style = style_fail
        if "-4" in str(excel_dict[n][m][1]):
            result = "TBD"
            ut_sheet.cell(row, col + 5, result).style = style_tbd

        if result == "IGNORED":
            t_ignored += 1
        elif result == "PASS":
            t_pass += 1
        elif result == "ERROR":
            t_error += 1
        elif result == "FAIL":
            t_fail += 1
        else:
            t_tbd += 1

# Initializing row and col to Zero for Summary sheet
row, col = 1, 1

summary.cell(row + 1, col + 1, file_extension + '_Results').style = style_header_s
summary.merge_cells(None, row + 1, col + 1, row + 1, col + 2)
summary.cell(row + 2, col + 1, "Total").style = style_title
summary.cell(row + 3, col + 1, "PASS").style = style_title
summary.cell(row + 4, col + 1, "FAIL").style = style_title
summary.cell(row + 5, col + 1, "ERROR").style = style_title
summary.cell(row + 6, col + 1, "IGNORED").style = style_title
summary.cell(row + 7, col + 1, "TBD").style = style_title

summary.cell(row + 2, col + 2, t_total).style = style_total
summary.cell(row + 3, col + 2, t_pass).style = style_pass
summary.cell(row + 4, col + 2, t_fail).style = style_fail
summary.cell(row + 5, col + 2, t_error).style = style_error
summary.cell(row + 6, col + 2, t_ignored).style = style_ignored
summary.cell(row + 7, col + 2, t_tbd).style = style_tbd

# Creating a bar chart
bar_chart_ut = BarChart()
bar_chart_ut.type = "col"
bar_chart_ut.style = 26
bar_chart_ut.title = file_extension + "_Results"
bar_chart_ut.y_axis.title = "Test_Cases"
bar_chart_ut.x_axis.title = "Execution"
bar_chart_ut.legend = None
bar_chart_ut.varyColors = True
bar_chart_ut.dataLabels = DataLabelList()
bar_chart_ut.dataLabels.showVal = True

bar3d_chart_ut = BarChart3D()
bar3d_chart_ut.type = "col"
bar3d_chart_ut.style = 26
bar3d_chart_ut.title = file_extension + "_Results"
bar3d_chart_ut.y_axis.title = "Test_Cases"
bar3d_chart_ut.x_axis.title = "Execution"
bar3d_chart_ut.legend = None
bar3d_chart_ut.varyColors = True
bar3d_chart_ut.dataLabels = DataLabelList()
bar3d_chart_ut.dataLabels.showVal = True

col = 1
row = 1

ut_chart_values = Reference(summary, col + 2, row + 1, col + 2, row + 7)
ut_chart_data_category = Reference(summary, col + 1, row + 2, col + 1, row + 7)
bar_chart_ut.add_data(ut_chart_values, from_rows=False, titles_from_data=True)
bar_chart_ut.set_categories(ut_chart_data_category)
summary.add_chart(bar_chart_ut, "F02")

bar3d_chart_ut.add_data(ut_chart_values, from_rows=False, titles_from_data=True)
bar3d_chart_ut.set_categories(ut_chart_data_category)
summary.add_chart(bar3d_chart_ut, "O02")

try:
    # Save the Workbook
    wb.save(save_file_path)
except PermissionError:
    print(">>>>>> Please close the {} sheet and Enter C to continue <<<<<<".format(result_file))
    restart = input().lower()
    while restart != "c":
        print(">>>>>> Please close the {} sheet and Enter C to continue <<<<<<".format(result_file))
        restart = input().lower()
    wb.save(save_file_path)

print(">>>>>> UT summary sheet has been created @ {} <<<<<<".format(save_file_path))
